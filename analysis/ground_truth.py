"""
Ground truth: run the validated engine over every real scan in DATA/ and join
the weighed masses from ``DATA/Yield Data Measurements.xlsx``.

Output rows are the target dataset every simulation front fits against:

- per-scan XRD metrics (DG%, Lc, crystalline fraction, norm_002_area, ...)
  via ``engine/research/trends.analyze_one`` — the same code path used on
  synthetic patterns, so sim-vs-real comparisons are apples-to-apples;
- per-run mass balance (mass yield, crystalline-graphite yield, wash QC)
  via ``engine/research/yield_calc.compute_from_name`` for the runs that have
  weighed pellet / post-furnace / post-acid masses.

Everything derived from DATA/ is confidential — write results only under
``outputs/`` (gitignored), never into committed files.
"""
from __future__ import annotations

import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
for _p in (_ROOT / "engine", _ROOT / "engine" / "research"):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

from run_parser import parse_run_filename  # noqa: E402
from trends import analyze_dir, make_plots, write_csv  # noqa: E402
from xrd_analyzer import XRDPattern, calibrate_internal_standard, fit_netl  # noqa: E402
from yield_calc import compute_from_name  # noqa: E402

DATA_DIR = _ROOT / "DATA"
MASSES_XLSX = DATA_DIR / "Yield Data Measurements.xlsx"
OUT_DIR = _ROOT / "outputs"


def internal_standard_dg(path: str) -> dict:
    """Re-fit DG% with an internal-standard 2theta correction.

    Residual Fe / Fe3C / CaO peaks act as an internal standard: their known
    line positions calibrate out specimen-displacement / zero-offset error that
    otherwise biases the (002) position — and DG% is ~1.4% per 0.01 deg, so a
    tenth-of-a-degree offset is the difference between a physical 93% and an
    impossible 107%. Applies the correction only when the offset is *significant*
    (>= 0.05 deg, the reference-lattice noise floor); otherwise DG is unchanged.
    """
    p = XRDPattern.from_file(path)
    tt, inten = p.two_theta, p.intensity
    raw = fit_netl(tt, inten, peak_count=2)
    cal = calibrate_internal_standard(tt, inten, phase="auto")
    applied = bool(cal.get("significant"))
    if applied:
        corr = fit_netl(tt, inten, peak_count=2, two_theta_offset=-cal["offset"])
        dg = corr["DG_percent"]
    else:
        dg = raw["DG_percent"]
    return {
        "DG_percent_raw": round(float(raw["DG_percent"]), 2),
        "DG_percent": round(float(dg), 2),
        "internal_standard": cal.get("phase"),
        "istd_offset_deg": cal.get("offset"),
        "istd_n_lines": cal.get("n_lines"),
        "istd_applied": applied,
    }


def recipe_key(name: str) -> tuple | None:
    """Canonical join key for a run: (C ratio, Fe ratio, CaCO3 ratio, T, t).
    Filenames and spreadsheet sample names both parse through run_parser, so
    naming-style differences (dashes vs underscores, 5hr vs 5H) collapse."""
    p = parse_run_filename(name)
    vals = (p.get("carbon_ratio"), p.get("fe_ratio"), p.get("caco3_ratio"),
            p.get("temperature_C"), p.get("time_h"))
    if any(v is None for v in vals):
        return None
    return tuple(round(float(v), 4) for v in vals)


def read_masses(xlsx: Path = MASSES_XLSX) -> list[dict]:
    """Rows of {sample, pellet, post_furnace, post_acid} from the lab sheet.
    Header row is 'Sample Name | Pellet | Post Furnace | Post Acid'."""
    if not xlsx.exists():
        return []
    from openpyxl import load_workbook
    ws = load_workbook(xlsx, read_only=True, data_only=True).active
    rows, header_seen = [], False
    for row in ws.iter_rows(values_only=True):
        cells = list(row) + [None] * (4 - len(row))
        if not header_seen:
            header_seen = str(cells[0]).strip().lower() == "sample name"
            continue
        sample = (str(cells[0]).strip() if cells[0] else "")
        if not sample:
            continue
        try:
            pellet = float(cells[1])
            post_furnace = float(cells[2])
        except (TypeError, ValueError):
            continue
        post_acid = None
        try:
            post_acid = float(cells[3])
        except (TypeError, ValueError):
            pass
        rows.append({"sample": sample, "pellet": pellet,
                     "post_furnace": post_furnace, "post_acid": post_acid})
    return rows


def build_ground_truth(data_dir: Path = DATA_DIR) -> dict:
    """Return {"metrics": [per-scan rows], "yields": [per-run mass rows]}.

    Every mass row gains the crystalline fraction measured on its matching scan
    (joined by recipe key), so its yield block already contains the corrected
    crystalline-graphite yield — the number the project is chasing.
    """
    metrics = analyze_dir(str(data_dir))
    for m in metrics:
        # "unlabeled = puck": every run in this dataset was pressed into a puck
        # unless the filename explicitly says powder, so fill the blank form.
        if not m.get("form"):
            m["form"] = "puck"
        if m.get("error") is None:
            try:
                istd = internal_standard_dg(str(data_dir / m["filename"]))
                m.update(istd)  # overrides DG_percent with the corrected value
            except (ValueError, FileNotFoundError):
                pass
    by_key: dict[tuple, dict] = {}
    for m in metrics:
        k = recipe_key(m["filename"])
        if k is not None and m.get("error") is None:
            # prefer the first clean scan per recipe (repeats share a key)
            by_key.setdefault(k, m)

    yields = []
    for mrow in read_masses():
        k = recipe_key(mrow["sample"])
        scan = by_key.get(k) if k else None
        cf = scan.get("crystalline_fraction") if scan else None
        try:
            y = compute_from_name(mrow["sample"], pellet=mrow["pellet"],
                                  post_furnace=mrow["post_furnace"],
                                  post_acid=mrow["post_acid"],
                                  crystalline_fraction=cf)
        except ValueError as exc:
            yields.append({"sample": mrow["sample"], "error": str(exc)})
            continue
        yields.append({"sample": mrow["sample"], "masses": mrow,
                       "matched_scan": scan["filename"] if scan else None,
                       "DG_percent": scan.get("DG_percent") if scan else None,
                       "result": y})
    return {"metrics": metrics, "yields": yields}


def _yield_csv_rows(yields: list[dict]) -> list[dict]:
    out = []
    for r in yields:
        if "error" in r:
            out.append({"sample": r["sample"], "error": r["error"]})
            continue
        res, y = r["result"], r["result"]["yield"]
        wash = res["wash"]["wash_check"] or {}
        out.append({
            "sample": r["sample"],
            "matched_scan": r["matched_scan"],
            "pellet": res["inputs"]["pellet"],
            "post_furnace": res["inputs"]["post_furnace"],
            "post_acid": res["inputs"]["post_acid"],
            "DG_percent": r["DG_percent"],
            "crystalline_fraction": y.get("crystalline_fraction"),
            "mass_yield_pct": y.get("mass_yield_pct"),
            "crystalline_graphite_yield_pct": y.get("crystalline_graphite_yield_pct"),
            "graphite_theoretical_g": res["chemistry"]["graphite_theoretical"],
            "measured_C_after_furnace_g": y.get("measured_C_after_furnace"),
            "boudouard_C_loss_g": res["chemistry"]["boudouard_C_loss"],
            "unaccounted_furnace_loss_g": res["reconciliation"]["unaccounted_furnace_loss"],
            "trapped_metal_g": wash.get("trapped_metal"),
            "wash_efficiency_pct": wash.get("wash_efficiency_pct"),
            "error": None,
        })
    return out


def main(out_dir: Path = OUT_DIR, plots: bool = True) -> dict:
    out_dir.mkdir(parents=True, exist_ok=True)
    gt = build_ground_truth()
    write_csv(gt["metrics"], str(out_dir / "ground_truth.csv"))

    import csv
    yrows = _yield_csv_rows(gt["yields"])
    if yrows:
        cols = list(yrows[0].keys())
        with open(out_dir / "yield_summary.csv", "w", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            w.writerows(yrows)

    made = make_plots(gt["metrics"], str(out_dir / "trend_plots")) if plots else []
    n_ok = sum(1 for m in gt["metrics"] if m.get("error") is None)
    print(f"scans analyzed : {n_ok}/{len(gt['metrics'])} clean")
    print(f"mass rows      : {len(gt['yields'])} "
          f"({sum(1 for r in gt['yields'] if r.get('matched_scan'))} matched to scans)")
    print(f"wrote          : {out_dir/'ground_truth.csv'}")
    if yrows:
        print(f"wrote          : {out_dir/'yield_summary.csv'}")
    for p in made:
        print(f"plot           : {p}")
    return gt


if __name__ == "__main__":
    main()
